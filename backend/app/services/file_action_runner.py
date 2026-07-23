from __future__ import annotations

import asyncio
import csv
import json
import shutil
import tarfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.core import storage
from app.models.schemas import ScrapeResult
from app.services.runtime_variables import RuntimeVariableStore, stringify_variable_value

type FlowNode = dict[str, object]

_FILE_ACTION_TYPES = {"file.read", "file.write", "file.copy", "file.move", "file.delete", "file.list", "file.step", "file.compress", "file.rename", "file.watch"}
_EXCEL_ACTION_TYPES = {"excel.read", "excel.write", "excel.step", "excel.addrow", "excel.save", "excel.deleterow", "excel.filter"}
# 单文件读取/CSV 解析上限：防止大文件把结果塞进变量/日志导致内存或前端渲染卡死。
_MAX_TEXT_BYTES = 512_000
_MAX_CSV_ROWS = 10_000


@dataclass(frozen=True)
class FileActionResult:
    action_type: str
    path: str
    values: list[str]
    count: int

    def to_scrape_result(self) -> ScrapeResult:
        return ScrapeResult(url=self.path, selector=self.action_type, count=self.count, values=self.values)


class FileActionRunner:
    def __init__(self, workspace_root: Path | None = None) -> None:
        self._workspace_root = (workspace_root or _resolve_workspace_root()).resolve()

    async def run(self, node: FlowNode, variables: RuntimeVariableStore, *, timeout_ms: int) -> FileActionResult:
        # file.watch 本身就是一个轮询等待循环，不能塞进 to_thread 的一次性同步调用里，需单独走异步轮询实现。
        if _read_action_type(node) == "file.watch":
            return await self._run_file_watch(node, variables, timeout_ms=timeout_ms)
        timeout_seconds = max(1, timeout_ms) / 1000
        # +1s 余量确保外层 wait_for 不会抢在线程内部逻辑之前触发，避免误判为超时。
        return await asyncio.wait_for(asyncio.to_thread(self._run_sync, node, variables), timeout=timeout_seconds + 1)

    def _run_sync(self, node: FlowNode, variables: RuntimeVariableStore) -> FileActionResult:
        action_type = _read_action_type(node)
        if action_type in _FILE_ACTION_TYPES:
            return self._run_file_action(action_type, node, variables)
        if action_type in _EXCEL_ACTION_TYPES:
            return self._run_excel_action(action_type, node, variables)
        raise ValueError(f"不支持的文件/Excel 节点类型: {action_type}")

    def _run_file_action(self, action_type: str, node: FlowNode, variables: RuntimeVariableStore) -> FileActionResult:
        if action_type == "file.step":
            action_type = "file.read"
        path = self._resolve_path(node, variables)
        if action_type == "file.read":
            if not path.exists() or not path.is_file():
                raise FileNotFoundError(f"文件不存在: {path.name}")
            content = _read_text(path)
            return FileActionResult(action_type=action_type, path=str(path), values=[content], count=1)

        if action_type == "file.write":
            content = variables.resolve_text(stringify_variable_value(node.get("content", node.get("value", ""))))
            _ensure_parent(path)
            path.write_text(content, encoding="utf-8")
            return FileActionResult(action_type=action_type, path=str(path), values=[str(path)], count=1)

        if action_type in {"file.copy", "file.move"}:
            target_path = self._resolve_target_path(node, variables)
            if not path.exists() or not path.is_file():
                raise FileNotFoundError(f"源文件不存在: {path.name}")
            _ensure_parent(target_path)
            if action_type == "file.copy":
                shutil.copy2(path, target_path)
            else:
                shutil.move(str(path), str(target_path))
            return FileActionResult(action_type=action_type, path=str(target_path), values=[str(target_path)], count=1)

        if action_type == "file.delete":
            if not path.exists():
                return FileActionResult(action_type=action_type, path=str(path), values=[], count=0)
            if path.is_dir():
                raise ValueError("file.delete 当前仅允许删除文件")
            path.unlink()
            return FileActionResult(action_type=action_type, path=str(path), values=[str(path)], count=1)

        if action_type == "file.list":
            if not path.exists() or not path.is_dir():
                raise FileNotFoundError(f"目录不存在: {path.name}")
            pattern = variables.resolve_text(_read_optional_string(node, "pattern") or "*")
            values = [str(item.relative_to(self._workspace_root)) for item in sorted(path.glob(pattern)) if item.is_file()]
            if len(values) > _MAX_CSV_ROWS:
                raise ValueError("目录遍历结果超过 10000 个文件")
            return FileActionResult(action_type=action_type, path=str(path), values=values, count=len(values))

        if action_type == "file.compress":
            target_path = self._resolve_target_path(node, variables)
            operation = _read_optional_string(node, "operation") or "compress"
            _ensure_parent(target_path)
            if operation == "decompress":
                if not path.exists():
                    raise FileNotFoundError(f"压缩文件不存在: {path.name}")
                if path.name.endswith(".tar.gz") or path.name.endswith(".tgz"):
                    with tarfile.open(path, "r:gz") as tf:
                        tf.extractall(str(target_path))
                else:
                    with zipfile.ZipFile(path, "r") as zf:
                        zf.extractall(str(target_path))
            else:
                if not path.exists():
                    raise FileNotFoundError(f"源路径不存在: {path.name}")
                if str(target_path).endswith(".tar.gz") or str(target_path).endswith(".tgz"):
                    with tarfile.open(target_path, "w:gz") as tf:
                        tf.add(path, arcname=path.name)
                else:
                    with zipfile.ZipFile(target_path, "w", zipfile.ZIP_DEFLATED) as zf:
                        if path.is_dir():
                            for item in sorted(path.rglob("*")):
                                if item.is_file():
                                    zf.write(item, item.relative_to(path.parent))
                        else:
                            zf.write(path, path.name)
            return FileActionResult(action_type=action_type, path=str(target_path), values=[str(target_path)], count=1)

        if action_type == "file.rename":
            target_path = self._resolve_target_path(node, variables)
            if not path.exists():
                raise FileNotFoundError(f"源文件不存在: {path.name}")
            _ensure_parent(target_path)
            shutil.move(str(path), str(target_path))
            return FileActionResult(action_type=action_type, path=str(target_path), values=[str(target_path)], count=1)

        raise ValueError(f"不支持的文件节点类型: {action_type}")

    def _run_excel_action(self, action_type: str, node: FlowNode, variables: RuntimeVariableStore) -> FileActionResult:
        if action_type == "excel.step":
            action_type = "excel.read"
        path = self._resolve_path(node, variables)

        if action_type == "excel.read":
            rows = _read_table_rows(path)
            values = _select_table_values(rows, node, variables)
            return FileActionResult(action_type=action_type, path=str(path), values=values, count=len(rows))

        if action_type == "excel.addrow":
            payload = _read_addrow_payload(node, variables)
            if _is_xlsx_path(path):
                count = _append_xlsx_row(path, payload, sheet_name=_read_optional_string(node, "sheetName"))
            else:
                count = _append_csv_row(path, payload)
            return FileActionResult(action_type=action_type, path=str(path), values=[str(path)], count=count)

        if action_type == "excel.save":
            # addrow/write 已同步落盘，这里只是登记最终行数，不重复写入。
            count = 0
            if path.exists():
                if _is_xlsx_path(path):
                    wb = load_workbook(path, read_only=True, data_only=True)
                    try:
                        ws = wb[_normalize_sheet_name(_read_optional_string(node, "sheetName"))] if _read_optional_string(node, "sheetName") in wb.sheetnames else wb.active
                        count = ws.max_row
                    finally:
                        wb.close()
                else:
                    with path.open("r", encoding="utf-8-sig", newline="") as f:
                        count = sum(1 for _ in csv.reader(f))
            return FileActionResult(action_type=action_type, path=str(path), values=[str(path)], count=count)

        if action_type == "excel.deleterow":
            if not path.exists():
                return FileActionResult(action_type=action_type, path=str(path), values=[str(path)], count=0)
            index = int(node.get("tabIndex", node.get("rowIndex", 1)))
            if _is_xlsx_path(path):
                count = _delete_xlsx_row(path, index, sheet_name=_read_optional_string(node, "sheetName"))
            else:
                with path.open("r", encoding="utf-8-sig", newline="") as f:
                    rows_list = list(csv.reader(f))
                if 0 <= index < len(rows_list):
                    rows_list.pop(index)
                with path.open("w", encoding="utf-8-sig", newline="") as f:
                    csv.writer(f).writerows(rows_list)
                count = len(rows_list)
            return FileActionResult(action_type=action_type, path=str(path), values=[str(path)], count=count)

        if action_type == "excel.filter":
            rows = _read_table_rows(path)
            column = _read_optional_string(node, "column")
            operation = str(node.get("operation", "filter"))
            pattern = _read_optional_string(node, "pattern") or ""
            if operation == "filter" and column:
                rows = [r for r in rows if r.get(column, "") == pattern]
            elif operation == "sort_asc" and column:
                rows = sorted(rows, key=lambda r: r.get(column, ""))
            elif operation == "sort_desc" and column:
                rows = sorted(rows, key=lambda r: r.get(column, ""), reverse=True)
            values = [json.dumps(r, ensure_ascii=False, separators=(",", ":")) for r in rows]
            return FileActionResult(action_type=action_type, path=str(path), values=values, count=len(rows))

        # excel.write：整文件覆盖写入
        rows = _read_rows_for_write(node, variables)
        _ensure_parent(path)
        if _is_xlsx_path(path):
            _write_xlsx_rows(path, rows, sheet_name=_read_optional_string(node, "sheetName"))
        else:
            with path.open("w", encoding="utf-8-sig", newline="") as file:
                writer = csv.writer(file)
                writer.writerows(rows)
        return FileActionResult(action_type=action_type, path=str(path), values=[str(path)], count=len(rows))

    async def _run_file_watch(self, node: FlowNode, variables: RuntimeVariableStore, *, timeout_ms: int) -> FileActionResult:
        path = self._resolve_path(node, variables)
        pattern = variables.resolve_text(_read_optional_string(node, "pattern") or "*")
        timeout_seconds = max(1, timeout_ms) / 1000

        initial_files: set[str] = set()
        initial_mtime: float | None = None
        if path.is_dir():
            initial_files = {str(f) for f in path.glob(pattern) if f.is_file()}
        elif path.exists():
            initial_mtime = path.stat().st_mtime

        loop = asyncio.get_event_loop()
        deadline = loop.time() + timeout_seconds
        while loop.time() < deadline:
            await asyncio.sleep(1.0)
            if path.is_dir():
                current_files = {str(f) for f in path.glob(pattern) if f.is_file()}
                new_files = sorted(current_files - initial_files)
                if new_files:
                    return FileActionResult(action_type="file.watch", path=str(path), values=new_files, count=len(new_files))
            elif path.exists():
                current_mtime = path.stat().st_mtime
                if initial_mtime is None or current_mtime != initial_mtime:
                    return FileActionResult(action_type="file.watch", path=str(path), values=[str(path)], count=1)
            else:
                pass

        raise TimeoutError(f"file.watch 超时：未检测到文件变化 ({path.name})")

    def _resolve_path(self, node: FlowNode, variables: RuntimeVariableStore) -> Path:
        raw_path = _read_required_string(node, "path", fallback_keys=("filePath", "targetPath", "targetUrl"))
        rendered = variables.resolve_text(raw_path)
        return self._resolve_relative_path(rendered)

    def _resolve_target_path(self, node: FlowNode, variables: RuntimeVariableStore) -> Path:
        raw_path = _read_required_string(node, "targetPath", fallback_keys=("destinationPath", "toPath"))
        rendered = variables.resolve_text(raw_path)
        return self._resolve_relative_path(rendered)

    def _resolve_relative_path(self, rendered: str) -> Path:
        # 安全边界：禁止绝对路径与 ".." 逃逸，流程节点只能读写 workspace_root 内的文件。
        path = Path(rendered)
        if path.is_absolute():
            raise ValueError("文件/Excel 节点只能使用相对路径")
        resolved = (self._workspace_root / path).resolve()
        if not resolved.is_relative_to(self._workspace_root):
            raise ValueError("文件路径超出 RPA 工作目录")
        return resolved


def is_file_action_node(node: FlowNode) -> bool:
    return node.get("type") in _FILE_ACTION_TYPES or node.get("type") in _EXCEL_ACTION_TYPES


def apply_file_result_variables(node: FlowNode, result: FileActionResult, variables: RuntimeVariableStore) -> list[str]:
    return _write_output_variables(node, variables, values=result.values, count=result.count)


def _write_output_variables(node: FlowNode, variables: RuntimeVariableStore, *, values: list[str], count: int) -> list[str]:
    saved_names: list[str] = []
    output_variable = _read_optional_string(node, "outputVariable") or _read_optional_string(node, "responseVariable") or _read_optional_string(node, "resultVariable")
    if output_variable is not None:
        variables.set(output_variable, values, scope="局部")
        saved_names.append(output_variable)

    first_value_variable = _read_optional_string(node, "firstValueVariable")
    if first_value_variable is not None:
        variables.set(first_value_variable, values[0] if values else "", scope="局部")
        saved_names.append(first_value_variable)

    count_variable = _read_optional_string(node, "countVariable") or _read_optional_string(node, "statusVariable")
    if count_variable is not None:
        variables.set(count_variable, count, scope="局部")
        saved_names.append(count_variable)
    return saved_names


def _read_action_type(node: FlowNode) -> str:
    value = node.get("type")
    if not isinstance(value, str) or not value.strip():
        raise ValueError("文件/Excel 节点缺少 type")
    return value.strip()


def _read_text(path: Path) -> str:
    if path.stat().st_size > _MAX_TEXT_BYTES:
        raise ValueError("文件读取超过 512KB 限制")
    return path.read_text(encoding="utf-8")


def _read_table_rows(path: Path) -> list[dict[str, str]]:
    if _is_xlsx_path(path):
        return _read_xlsx_rows(path)
    return _read_csv_rows(path)


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"CSV 文件不存在: {path.name}")
    if path.suffix.lower() not in {".csv", ".txt"}:
        raise ValueError("Excel 节点读取仅支持 .csv / .xlsx")
    if path.stat().st_size > _MAX_TEXT_BYTES:
        raise ValueError("CSV 文件超过 512KB 限制")

    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        rows = [dict(row) for index, row in enumerate(reader) if index < _MAX_CSV_ROWS]
    if len(rows) >= _MAX_CSV_ROWS:
        raise ValueError("CSV 行数超过 10000 行限制")
    return rows


def _read_xlsx_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在: {path.name}")
    if path.stat().st_size > 10 * 1024 * 1024:
        raise ValueError("Excel 文件超过 10MB 限制")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not values:
        return []
    headers = [stringify_variable_value(cell) for cell in values[0]]
    rows: list[dict[str, str]] = []
    for row in values[1 : _MAX_CSV_ROWS + 1]:
        rows.append({header: stringify_variable_value(row[index] if index < len(row) else "") for index, header in enumerate(headers)})
    if len(rows) >= _MAX_CSV_ROWS:
        raise ValueError("Excel 行数超过 10000 行限制")
    return rows


def _select_table_values(rows: list[dict[str, str]], node: FlowNode, variables: RuntimeVariableStore) -> list[str]:
    column = _read_optional_string(node, "column")
    if column is not None:
        resolved_column = variables.resolve_text(column)
        return [row.get(resolved_column, "") for row in rows]
    return [json.dumps(row, ensure_ascii=False, separators=(",", ":")) for row in rows]


def _read_rows_for_write(node: FlowNode, variables: RuntimeVariableStore) -> list[list[str]]:
    raw_rows = node.get("rows", node.get("value", node.get("content", [])))
    if isinstance(raw_rows, str):
        rendered = variables.resolve_text(raw_rows)
        try:
            decoded = json.loads(rendered)
        except json.JSONDecodeError:
            return [[rendered]]
        raw_rows = decoded
    if not isinstance(raw_rows, list):
        raise ValueError("excel.write 节点 rows 必须是数组")

    dict_rows = [row for row in raw_rows if isinstance(row, dict)]
    if dict_rows and len(dict_rows) == len(raw_rows):
        headers: list[str] = []
        for row in dict_rows:
            for key in row.keys():
                header = str(key)
                if header not in headers:
                    headers.append(header)
        rows = [headers]
        rows.extend([[stringify_variable_value(row.get(header, "")) for header in headers] for row in dict_rows])
        if len(rows) > _MAX_CSV_ROWS:
            raise ValueError("CSV 写入不能超过 10000 行")
        return rows

    rows: list[list[str]] = []
    for raw_row in raw_rows:
        if isinstance(raw_row, dict):
            rows.append([stringify_variable_value(value) for value in raw_row.values()])
        elif isinstance(raw_row, list):
            rows.append([stringify_variable_value(value) for value in raw_row])
        else:
            rows.append([stringify_variable_value(raw_row)])
    if len(rows) > _MAX_CSV_ROWS:
        raise ValueError("CSV 写入不能超过 10000 行")
    return rows


def _read_addrow_payload(node: FlowNode, variables: RuntimeVariableStore) -> dict[str, str] | list[str]:
    raw = node.get("rowData", node.get("row", node.get("content", [])))
    if isinstance(raw, list):
        return [variables.resolve_text(str(cell)) if isinstance(cell, str) else str(cell) for cell in raw]
    if isinstance(raw, str):
        rendered = variables.resolve_text(raw)
        python_literal = _parse_python_literal(rendered)
        if isinstance(python_literal, dict):
            return {str(key): stringify_variable_value(value) for key, value in python_literal.items()}
        if isinstance(python_literal, list):
            return [stringify_variable_value(value) for value in python_literal]
        try:
            decoded = json.loads(rendered)
            if isinstance(decoded, dict):
                return {str(key): stringify_variable_value(value) for key, value in decoded.items()}
            if isinstance(decoded, list):
                return [stringify_variable_value(value) for value in decoded]
            return [stringify_variable_value(decoded)]
        except json.JSONDecodeError:
            return [rendered]
    if isinstance(raw, dict):
        return {str(key): stringify_variable_value(value) for key, value in raw.items()}
    return []


def _append_csv_row(path: Path, payload: dict[str, str] | list[str]) -> int:
    existing: list[list[str]] = []
    if path.exists() and path.stat().st_size > 0:
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            existing = list(csv.reader(file))

    if isinstance(payload, dict):
        if existing:
            headers = existing[0]
        else:
            headers = list(payload.keys())
            existing.append(headers)
        existing.append([payload.get(header, "") for header in headers])
    else:
        existing.append(payload)

    _ensure_parent(path)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        csv.writer(file).writerows(existing)
    return len(existing)


def _append_xlsx_row(path: Path, payload: dict[str, str] | list[str], *, sheet_name: str | None) -> int:
    _ensure_parent(path)
    workbook = load_workbook(path) if path.exists() and path.stat().st_size > 0 else Workbook()
    try:
        sheet = _select_or_create_sheet(workbook, sheet_name)
        is_blank = _is_blank_sheet(sheet)
        if isinstance(payload, dict):
            headers = _read_xlsx_headers(sheet)
            if not headers:
                headers = list(payload.keys())
                _write_xlsx_row(sheet, 1, headers)
                _write_xlsx_row(sheet, 2, [payload.get(header, "") for header in headers])
            else:
                sheet.append([payload.get(header, "") for header in headers])
        else:
            if is_blank:
                _write_xlsx_row(sheet, 1, payload)
            else:
                sheet.append(payload)
        workbook.save(path)
        return sheet.max_row
    finally:
        workbook.close()


def _write_xlsx_rows(path: Path, rows: list[list[str]], *, sheet_name: str | None) -> None:
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = _normalize_sheet_name(sheet_name)
        for row in rows:
            sheet.append(row)
        workbook.save(path)
    finally:
        workbook.close()


def _delete_xlsx_row(path: Path, index: int, *, sheet_name: str | None) -> int:
    workbook = load_workbook(path)
    try:
        sheet = _select_or_create_sheet(workbook, sheet_name)
        if 0 <= index < sheet.max_row:
            sheet.delete_rows(index + 1)
        workbook.save(path)
        return sheet.max_row
    finally:
        workbook.close()


def _read_xlsx_headers(sheet: Any) -> list[str]:
    if _is_blank_sheet(sheet):
        return []
    if sheet.max_row < 1:
        return []
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [stringify_variable_value(cell) for cell in first_row]
    return [header for header in headers if header]


def _select_or_create_sheet(workbook: Any, sheet_name: str | None) -> Any:
    normalized = _normalize_sheet_name(sheet_name)
    if normalized in workbook.sheetnames:
        return workbook[normalized]
    sheet = workbook.active
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
        sheet.title = normalized
        return sheet
    return workbook.create_sheet(normalized)


def _is_blank_sheet(sheet: Any) -> bool:
    return sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None


def _write_xlsx_row(sheet: Any, row_index: int, values: list[str]) -> None:
    for column_index, value in enumerate(values, start=1):
        sheet.cell(row=row_index, column=column_index, value=value)


def _normalize_sheet_name(value: str | None) -> str:
    name = (value or "Sheet1").strip()[:31]
    for char in "[]:*?/\\": 
        name = name.replace(char, "_")
    return name or "Sheet1"


def _is_xlsx_path(path: Path) -> bool:
    return path.suffix.lower() in {".xlsx", ".xlsm"}


def _parse_python_literal(value: str) -> object:
    stripped = value.strip()
    if not stripped or stripped[0] not in "{[":
        return None
    try:
        import ast

        return ast.literal_eval(stripped)
    except (SyntaxError, ValueError):
        return None


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _resolve_workspace_root() -> Path:
    return storage.resolve_workspace_root()


def _read_required_string(node: FlowNode, key: str, *, fallback_keys: tuple[str, ...] = ()) -> str:
    for candidate_key in (key, *fallback_keys):
        value = node.get(candidate_key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    raise ValueError(f"文件/Excel 节点缺少 {key}")


def _read_optional_string(node: FlowNode, key: str) -> str | None:
    value = node.get(key)
    return value.strip() if isinstance(value, str) and value.strip() else None
