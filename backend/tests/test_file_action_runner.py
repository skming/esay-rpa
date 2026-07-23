from __future__ import annotations

from openpyxl import load_workbook

from app.services.file_action_runner import FileActionRunner, apply_file_result_variables
from app.services.runtime_variables import RuntimeVariableStore


async def test_excel_read_csv_writes_output_variables(tmp_path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    (workspace / "orders.csv").write_text("order_id,total\nA001,42\nA002,64\n", encoding="utf-8")
    runner = FileActionRunner(workspace)
    variables = RuntimeVariableStore.from_initial({})

    node = {
        "type": "excel.read",
        "path": "orders.csv",
        "column": "order_id",
        "outputVariable": "order_ids",
        "firstValueVariable": "first_order_id",
        "countVariable": "row_count",
    }
    result = await runner.run(
        node,
        variables,
        timeout_ms=1000,
    )
    apply_file_result_variables(node, result, variables)

    assert result.count == 2
    assert result.values == ["A001", "A002"]
    snapshots = {variable.name: variable for variable in variables.snapshots()}
    assert snapshots["order_ids"].value == '["A001", "A002"]'
    assert snapshots["first_order_id"].value == "A001"
    assert snapshots["row_count"].value == "2"


async def test_excel_read_csv_accepts_response_and_status_variable_aliases(tmp_path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    (workspace / "orders.csv").write_text("order_id,total\nA001,42\nA002,64\n", encoding="utf-8")
    runner = FileActionRunner(workspace)
    variables = RuntimeVariableStore.from_initial({})
    node = {
        "type": "excel.read",
        "path": "orders.csv",
        "column": "order_id",
        "responseVariable": "order_ids",
        "statusVariable": "row_count",
    }

    result = await runner.run(node, variables, timeout_ms=1000)
    saved = apply_file_result_variables(node, result, variables)

    assert saved == ["order_ids", "row_count"]
    assert variables.get("order_ids") == ["A001", "A002"]
    assert variables.get("row_count") == 2


async def test_excel_write_dict_rows_writes_headers_and_aligned_values(tmp_path) -> None:
    runner = FileActionRunner(tmp_path)
    variables = RuntimeVariableStore.from_initial(
        {
            "rows": [
                {"序号": "1", "合约编号": "Y001", "门店名称": "门店A"},
                {"序号": "2", "合约编号": "Y002", "门店名称": "门店B"},
            ]
        }
    )

    await runner.run({"type": "excel.write", "path": "contracts.xlsx", "rows": "${var.rows}"}, variables, timeout_ms=1000)

    workbook = load_workbook(tmp_path / "contracts.xlsx", read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    assert rows == [
        ("序号", "合约编号", "门店名称"),
        ("1", "Y001", "门店A"),
        ("2", "Y002", "门店B"),
    ]


async def test_file_write_rejects_path_traversal(tmp_path) -> None:
    runner = FileActionRunner(tmp_path)
    variables = RuntimeVariableStore.from_initial({})

    try:
        await runner.run({"type": "file.write", "path": "../escape.txt", "content": "unsafe"}, variables, timeout_ms=1000)
    except ValueError as exc:
        assert "超出" in str(exc)
    else:
        raise AssertionError("文件节点必须拒绝目录穿越路径")


async def test_file_read_resolves_template_path(tmp_path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    (workspace / "report.txt").write_text("done", encoding="utf-8")
    runner = FileActionRunner(workspace)
    variables = RuntimeVariableStore.from_initial({"name": "report"})

    node = {"type": "file.read", "path": "${var.name}.txt", "outputVariable": "content"}
    result = await runner.run(node, variables, timeout_ms=1000)
    apply_file_result_variables(node, result, variables)

    assert result.values == ["done"]
    snapshots = {variable.name: variable for variable in variables.snapshots()}
    assert snapshots["content"].value == '["done"]'


async def test_file_copy_move_delete_and_list(tmp_path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    (workspace / "source.txt").write_text("hello", encoding="utf-8")
    (workspace / "keep.log").write_text("skip", encoding="utf-8")
    runner = FileActionRunner(workspace)
    variables = RuntimeVariableStore.from_initial({})

    copy_result = await runner.run({"type": "file.copy", "path": "source.txt", "targetPath": "archive/source.txt"}, variables, timeout_ms=1000)
    assert copy_result.values == [str(workspace / "archive/source.txt")]
    assert (workspace / "archive/source.txt").read_text(encoding="utf-8") == "hello"

    move_result = await runner.run({"type": "file.move", "path": "archive/source.txt", "targetPath": "done/source.txt"}, variables, timeout_ms=1000)
    assert move_result.values == [str(workspace / "done/source.txt")]
    assert not (workspace / "archive/source.txt").exists()
    assert (workspace / "done/source.txt").exists()

    list_result = await runner.run({"type": "file.list", "path": ".", "pattern": "*.txt", "outputVariable": "files", "countVariable": "file_count"}, variables, timeout_ms=1000)
    apply_file_result_variables({"outputVariable": "files", "countVariable": "file_count"}, list_result, variables)
    assert list_result.values == ["source.txt"]
    assert variables.get("file_count") == 1

    delete_result = await runner.run({"type": "file.delete", "path": "source.txt"}, variables, timeout_ms=1000)
    assert delete_result.count == 1
    assert not (workspace / "source.txt").exists()
